import json

import pandas as pd
from django.shortcuts import render
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from selenium.webdriver.support import expected_conditions as EC
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from openpyxl import load_workbook
import random
import openpyxl

from .models import *


@csrf_exempt
@require_http_methods(['POST'])
def update_password(request):
    data = json.loads(request.body)
    new_password = data.get('password')
    static_data = StaticData.objects.get(static_id=1)
    static_data.static_password = new_password
    static_data.save()
    return JsonResponse({'msg': '密码修改成功'})


@csrf_exempt
@require_http_methods(['POST'])
def get_password(request):
    static_data = StaticData.objects.get(static_id=1)
    return JsonResponse({'password': static_data.static_password})


@csrf_exempt
@require_http_methods(['POST'])
def update_cookie(request):
    data = json.loads(request.body)
    new_cookie = data.get('cookie')
    new_lvt = data.get('lvt')
    new_lpvt = data.get('lpvt')
    static_data = StaticData.objects.get(static_id=1)
    static_data.static_cookie = new_cookie
    static_data.static_lvt = new_lvt
    static_data.static_lpvt = new_lpvt
    static_data.save()
    return JsonResponse({'msg': 'cookie修改成功'})


@csrf_exempt
@require_http_methods(['POST'])
def get_cookie(request):
    static_data = StaticData.objects.get(static_id=1)
    return JsonResponse({'cookie': static_data.static_cookie})


@csrf_exempt
@require_http_methods(['POST'])
def create_class(request):
    data = json.loads(request.body)
    class_name = data.get('name')
    class_teacher = data.get('teacher')
    class_year = data.get('year')
    class_season = data.get('season')
    new_class = StudentClass.objects.create(class_name=class_name, class_teacher=class_teacher, class_year=class_year, class_season=class_season)
    new_class.save()
    return JsonResponse({'class_id': new_class.class_id, 'msg': "创建班级成功"})


@csrf_exempt
@require_http_methods(['POST'])
def delete_class(request):
    data = json.loads(request.body)
    class_id = data.get('id')
    try:
        student_class = StudentClass.objects.get(class_id=class_id)
        student_class.delete()
        return JsonResponse({'msg': '班级删除成功'})
    except StudentClass.DoesNotExist:
        return JsonResponse({'msg': '班级不存在'})


@csrf_exempt
@require_http_methods(['POST'])
def update_class(request):
    data = json.loads(request.body)
    class_id = data.get('id')
    student_class = StudentClass.objects.get(class_id=class_id)
    class_name = data.get('name')
    class_teacher = data.get('teacher')
    class_year = data.get('year')
    class_season = data.get('season')
    student_class.class_name = class_name
    student_class.class_teacher = class_teacher
    student_class.class_year = class_year
    student_class.class_season = class_season
    student_class.save()
    return JsonResponse({'msg': '班级信息修改成功'})


@csrf_exempt
@require_http_methods(['POST'])
def get_classes(request):
    data = json.loads(request.body)
    class_year = data.get('year')
    class_season = data.get('season')
    student_classes = StudentClass.objects.filter(class_year=class_year, class_season=class_season)
    class_list = []
    for student_class in student_classes:
        class_list.append({
            'class_id': student_class.class_id,
            'name': student_class.class_name,
            'teacher': student_class.class_teacher,
            'year': student_class.class_year,
            'season': student_class.class_season,
            'students': len(student_class.student_list.all())
        })
    return JsonResponse({'classes': class_list})


@csrf_exempt
@require_http_methods(['POST'])
def get_all_classes(request):
    student_classes = StudentClass.objects.all()
    class_list = []
    for student_class in student_classes:
        class_list.append({
            'class_id': student_class.class_id,
            'name': student_class.class_name,
            'teacher': student_class.class_teacher,
            'year': student_class.class_year,
            'season': student_class.class_season,
            'students': len(student_class.student_list.all())
        })
    return JsonResponse({'classes': class_list})


@csrf_exempt
@require_http_methods(['POST'])
def get_class_student(request):
    data = json.loads(request.body)
    class_id = data.get('id')
    student_class = StudentClass.objects.get(class_id=class_id)
    student_list = student_class.student_list.all()
    students = []
    for student in student_list:
        students.append({
            'stu_id': student.stu_id,
            'stu_name': student.stu_name
        })
    return JsonResponse({'students': students})


@csrf_exempt
@require_http_methods(['POST'])
def add_student_to_class(request):
    data = json.loads(request.body)
    student_id = data.get('student_id')
    student_name = data.get('student_name')
    class_id = data.get('class_id')

    try:
        student_class = StudentClass.objects.get(class_id=class_id)
    except StudentClass.DoesNotExist:
        return JsonResponse({'msg': '班级不存在'})

    student = Student.objects.create(stu_id=student_id)
    student.stu_name = student_name
    student.class_id = class_id
    student.save()

    student_class.student_list.add(student)

    return JsonResponse({'msg': '学生已成功加入班级'})


@csrf_exempt
@require_http_methods(['POST'])
def add_students_list_to_class(request):
    class_id = request.POST.get('class_id')
    file = request.FILES.get('file')

    try:
        student_class = StudentClass.objects.get(class_id=class_id)
    except StudentClass.DoesNotExist:
        return JsonResponse({'msg': '班级不存在'})

    students_added = 0
    try:
        workbook = load_workbook(file)
        sheet = workbook.active
        for row in sheet.iter_rows(values_only=True):
            student_id, student_name = row[:2]
            student = Student.objects.create(stu_id=student_id)
            student.stu_name = student_name
            student.class_id = class_id
            student.save()
            student_class.student_list.add(student)
            students_added += 1

        return JsonResponse({'msg': f'成功添加 {students_added} 个学生到班级'})
    except Exception as e:
        return JsonResponse({'msg': f'添加学生失败：{str(e)}'})


@csrf_exempt
@require_http_methods(['POST'])
def export_students_list_in_class(request):
    data = json.loads(request.body)
    class_id = data.get('class_id')
    student_class = StudentClass.objects.get(class_id=class_id)

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    students = student_class.student_list.all()
    row = 1
    for student in students:
        sheet.cell(row=row, column=1, value=student.stu_id)
        sheet.cell(row=row, column=2, value=student.stu_name)
        row += 1

    # 创建响应对象
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=student_list_class_{class_id}.xlsx'

    # 保存文件到响应对象
    workbook.save(response)

    return response


@csrf_exempt
@require_http_methods(['POST'])
def delete_student(request):
    data = json.loads(request.body)
    stu_id = data.get('student_id')
    stu_name = data.get('student_name')
    class_id = data.get('class_id')
    student = Student.objects.get(stu_id=stu_id, stu_name=stu_name, class_id=class_id)
    student.delete()
    return JsonResponse({'msg': '学生删除成功'})


@csrf_exempt
@require_http_methods(['POST'])
def create_exam(request):
    data = json.loads(request.body)
    join_class = data.get('join_class')
    exam_name = data.get('name')
    exam_time = data.get('time')
    num = data.get('num')
    exam = Exam.objects.create(exam_name=exam_name, exam_time=exam_time)
    students = []
    for class_id in join_class:
        try:
            student_class = StudentClass.objects.get(class_id=class_id)
            students.extend(student_class.student_list.all())
            exam.join_class_list.add(student_class)
        except StudentClass.DoesNotExist:
            return JsonResponse({'msg': '班级不存在'})

    random.shuffle(students)

    total_students = len(students)
    students_per_room = total_students // num + 1
    for i in range(num):
        room_name = f'Room {i + 1}'
        room = ExamRoom.objects.create(er_name=room_name)
        for j in range(students_per_room):
            if i * students_per_room + j >= len(students):
                break
            student = students[i * students_per_room + j]
            student_table = StudentTable.objects.create(stu_id=student.stu_id, stu_name=student.stu_name)
            student_table.stu_room_num = i + 1
            student_table.stu_seat_num = j + 1
            student_table.save()
            room.er_student_list.add(student_table)
        room.save()
        exam.exam_room_list.add(room)
    exam.save()
    return JsonResponse({'msg': '考试和考场创建成功'})


@csrf_exempt
@require_http_methods(['POST'])
def get_all_exams(request):
    exams = Exam.objects.all()
    exam_list = []
    for exam in exams:
        class_list = []
        classes = exam.join_class_list.all()
        for now_class in classes:
            class_list.append(now_class.class_id)
        exam_list.append({
            'exam_id': exam.exam_id,
            'name': exam.exam_name,
            'time': exam.exam_time,
            'room_num': len(exam.exam_room_list.all()),
            'all_class_id': class_list
        })
    return JsonResponse({'exam_list': exam_list})


@csrf_exempt
@require_http_methods(['POST'])
def get_room_in_exam(request):
    data = json.loads(request.body)
    exam_id = data.get('id')
    exam = Exam.objects.get(exam_id=exam_id)
    room_list = exam.exam_room_list.all()
    rooms = []
    for room in room_list:
        rooms.append({
            'room_id': room.er_id,
            'room_name': room.er_name,
            'room_case_id': room.er_case_id,
            'room_stu_num': len(room.er_student_list.all())
        })
    return JsonResponse({'rooms': rooms})


@csrf_exempt
@require_http_methods(['POST'])
def get_student_info_in_room(request):
    data = json.loads(request.body)
    room_id = data.get('id')
    room = ExamRoom.objects.get(er_id=room_id)
    room_case = ExamRoomCase.objects.get(erc_id=room.er_case_id)
    stu_info_list = room.er_student_list.all()
    stu_infos = []
    for stu_info in stu_info_list:
        seat = room_case.erc_seats.get(ersc_seat_num=stu_info.stu_seat_num)
        stu_infos.append({
            'stu_id': stu_info.stu_id,
            'stu_name': stu_info.stu_name,
            'room_num': stu_info.stu_room_num,
            'seat_num': stu_info.stu_seat_num,
            'seat_x': seat.ersc_x,
            'seat_y': seat.ersc_y,
            'is_register': stu_info.is_register,
            'is_submit': stu_info.is_submit
        })
    return JsonResponse({'stu_infos': stu_infos})


@csrf_exempt
@require_http_methods(['POST'])
def get_class_in_exam(request):
    data = json.loads(request.body)
    exam_id = data.get('id')
    exam = Exam.objects.get(exam_id=exam_id)
    class_list = exam.join_class_list.all()
    classes = []
    for now_class in class_list:
        classes.append({
            'class_id': now_class.class_id,
            'name': now_class.class_name,
            'teacher': now_class.class_teacher,
            'year': now_class.class_year,
            'season': now_class.class_season,
            'students': len(now_class.student_list.all())
        })
    return JsonResponse({'classes': classes})


@csrf_exempt
@require_http_methods(['POST'])
def delete_exam(request):
    data = json.loads(request.body)
    exam_id = data.get('id')
    exam = Exam.objects.get(exam_id=exam_id)
    exam.delete()
    return JsonResponse({'msg': '考试删除成功'})


@csrf_exempt
@require_http_methods(['POST'])
def set_case_to_exam(request):
    data = json.loads(request.body)
    room_id = data.get('room_id')
    case_id = data.get('case_id')
    room = ExamRoom.objects.get(er_id=room_id)
    room.er_case_id = case_id
    room.save()
    return JsonResponse({'msg': '修改成功'})


@csrf_exempt
@require_http_methods(['POST'])
def create_class_case(request):
    data = json.loads(request.body)
    case_name = data.get('name')
    case_x_length = data.get('x')
    case_y_length = data.get('y')
    seats = data.get('seat_case')
    case = ExamRoomCase.objects.create(erc_name=case_name, erc_x_length=case_x_length, erc_y_length=case_y_length)
    for seat in seats:
        seat_case = ExamRoomSeatCase.objects.create(ersc_seat_num=seat['num'], ersc_x=seat['x'], ersc_y=seat['y'])
        seat_case.save()
        case.erc_seats.add(seat_case)
    case.save()
    return JsonResponse({'msg': '模型创建成功'})


@csrf_exempt
@require_http_methods(['POST'])
def get_all_class_case(request):
    case_list = ExamRoomCase.objects.all()
    cases = []
    for case in case_list:
        cases.append({
            'case_id': case.erc_id,
            'name': case.erc_name,
            'x_length': case.erc_x_length,
            'y_length': case.erc_y_length,
            'seat_num': len(case.erc_seats.all())
        })
    return JsonResponse({'all_cases': cases})


@csrf_exempt
@require_http_methods(['POST'])
def get_seats_in_case(request):
    data = json.loads(request.body)
    case_id = data.get('id')
    case = ExamRoomCase.objects.get(erc_id=case_id)
    seat_list = case.erc_seats.all()
    seats = []
    for seat in seat_list:
        seats.append({
            'seat_id': seat.ersc_id,
            'number': seat.ersc_seat_num,
            'x': seat.ersc_x,
            'y': seat.ersc_y
        })
    return JsonResponse({'seats': seats})


@csrf_exempt
@require_http_methods(['POST'])
def delete_case(request):
    data = json.loads(request.body)
    case_id = data.get('id')
    case = ExamRoomCase.objects.get(erc_id=case_id)
    case.delete()
    return JsonResponse({'msg': '考场模型删除成功'})


@csrf_exempt
@require_http_methods(['POST'])
def student_register(request):
    data = json.loads(request.body)
    room_id = data.get('room_id')
    seat_num = data.get('seat_num')
    is_register = data.get('is_register')
    exam_room = ExamRoom.objects.get(er_id=room_id)
    student_table = exam_room.er_student_list.get(stu_seat_num=seat_num)
    student_table.is_register = is_register
    student_table.save()
    return JsonResponse({'msg': f'学生{student_table.stu_id}的签到状态已改为{is_register}'})


@csrf_exempt
@require_http_methods(['POST'])
def upload_excel(request):
    excel_file = request.FILES['file']
    # 使用pandas读取Excel文件
    df = pd.read_excel(excel_file)

    # 遍历DataFrame的每一行，将数据存入数据库
    for index, row in df.iterrows():
        student = StudentTable(
            stu_id=row['stu_id'],
            stu_name=row['stu_name'],
            stu_room_num=row['stu_room_num'],
            stu_seat_num=row['stu_seat_num']
        )
        student.save()

    return HttpResponse('上传成功！')


@csrf_exempt
@require_http_methods(['POST'])
def clear_database(request):
    # 清空StudentTable表中的所有数据
    StudentTable.objects.all().delete()
    return HttpResponse('数据库已清空！')


@csrf_exempt
@require_http_methods(['POST'])
def get_information(request):
    data = json.loads(request.body)
    exam_id_list = data.get('exam_id_list')
    er_id_list = data.get('er_id_list')
    stu_table_list = []
    stu_er_list = []
    for er_id in er_id_list:
        er = ExamRoom.objects.get(er_id=er_id)
        for stu_table in er.er_student_list.all():
            stu_table_list.append(stu_table)
            stu_er_list.append(er_id)
    # 初始化Chrome浏览器
    options = webdriver.ChromeOptions()
    options.add_argument('--ignore-certificate-errors')
    driver = webdriver.Chrome(options=options)
    message_list = []
    static_data = StaticData.objects.get(static_id=1)
    cookie_data = {
        'JSESSIONID': static_data.static_cookie,
        'Hm_lvt_9eca16a516f8b449709378fbcbb6b200': static_data.static_lvt,
        'Hm_lpvt_9eca16a516f8b449709378fbcbb6b200': static_data.static_lpvt
    }
    for e_id in exam_id_list:
        driver.get(
            f"https://judge.buaa.edu.cn/admin/courseAdmin/examAdmin/examResult.jsp?examID={e_id}&courseFlag=1654")

        for name, value in cookie_data.items():
            cookie = {
                'name': name,
                'value': value
            }
            driver.add_cookie(cookie)
        driver.refresh()
        # element = driver.find_element(By.CLASS_NAME,"even")
        elements1 = WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.CLASS_NAME, "odd")))
        elements2 = WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.CLASS_NAME, "even")))
        elements = elements1 + elements2

        # 获取元素的文本内容
        for element in elements:
            text = element.text
            parts = text.split()
            print(parts[1], parts[2], parts[-3])
            room_num = -1
            seat_num = -1
            for index, stu_table in enumerate(stu_table_list):
                if stu_table.stu_id == parts[1]:
                    room_num = stu_er_list[index]
                    seat_num = stu_table.stu_seat_num
                    break
            message_info = {
                "stu_id": parts[1],
                "stu_name": parts[2],
                "score": parts[-3],
                "room_num": room_num,
                "seat_num": seat_num
            }
            message_list.append(message_info)
    # 关闭浏览器
    driver.quit()
    return JsonResponse({"submit_message": message_list})
